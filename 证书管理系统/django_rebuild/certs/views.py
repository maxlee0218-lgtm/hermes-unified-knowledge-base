import json
from datetime import timedelta
from io import BytesIO

import openpyxl
from django.db.models import Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods, require_POST
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import Certificate, Employee, Notification, QueryExecution, QueryTask
from .services import build_dashboard_context, process_pending_tasks_inline


def _parse_date(value):
    if not value:
        return None
    try:
        return timezone.datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("日期格式必须为 YYYY-MM-DD") from exc


def _request_data(request):
    if request.content_type and "application/json" in request.content_type:
        return json.loads(request.body.decode("utf-8") or "{}")
    if request.method in {"POST", "PUT", "PATCH"}:
        return request.POST.dict() if request.POST else json.loads(request.body.decode("utf-8") or "{}")
    return {}


def _employee_dict(employee):
    return {
        "id": employee.id,
        "name": employee.name,
        "id_number": employee.id_number,
        "department": employee.department,
        "phone": employee.phone,
        "position": employee.position,
        "employment_status": employee.employment_status,
        "created_at": employee.created_at.strftime("%Y-%m-%d"),
        "updated_at": employee.updated_at.strftime("%Y-%m-%d"),
    }


def _certificate_dict(cert):
    return {
        "id": cert.id,
        "employee_id": cert.employee_id,
        "employee_name": cert.employee.name,
        "id_number": cert.employee.id_number,
        "certificate_type": cert.certificate_type,
        "operation_item": cert.operation_item,
        "cert_number": cert.cert_number,
        "issuing_authority": cert.issuing_authority,
        "issue_date": cert.issue_date.isoformat() if cert.issue_date else "",
        "expiry_date": cert.expiry_date.isoformat() if cert.expiry_date else "",
        "review_date": cert.review_date.isoformat() if cert.review_date else "",
        "status": cert.status,
        "display_status": cert.display_status,
        "days_until_expiry": cert.days_until_expiry,
        "created_at": cert.created_at.strftime("%Y-%m-%d"),
    }


def _normalize_employee_payload(data, require_full=True):
    name = str(data.get("name", "")).strip()
    id_number = str(data.get("id_number") or data.get("id_card") or "",).strip()
    department = str(data.get("department", "")).strip()
    phone = str(data.get("phone", "")).strip()
    position = str(data.get("position", "")).strip()
    if require_full and (not name or not id_number):
        raise ValueError("姓名和身份证号为必填项")
    if id_number and len(id_number) != 18:
        raise ValueError("身份证号格式不正确（应为18位）")
    return {
        "name": name,
        "id_number": id_number,
        "department": department,
        "phone": phone,
        "position": position,
    }


def _normalize_certificate_payload(data):
    employee_id = data.get("employee_id")
    if not employee_id:
        raise ValueError("请选择员工")
    cert_type = str(data.get("certificate_type", "")).strip()
    if not cert_type:
        raise ValueError("证书类型为必填项")
    return {
        "employee_id": int(employee_id),
        "certificate_type": cert_type,
        "operation_item": str(data.get("operation_item", "")).strip(),
        "issuing_authority": str(data.get("issuing_authority", "")).strip(),
        "cert_number": str(data.get("cert_number", "")).strip(),
        "issue_date": _parse_date(str(data.get("issue_date", "")).strip()),
        "expiry_date": _parse_date(str(data.get("expiry_date", "")).strip()),
        "review_date": _parse_date(str(data.get("review_date", "")).strip()),
        "status": str(data.get("status", Certificate.Status.VALID)).strip() or Certificate.Status.VALID,
    }


@require_GET
def health(request):
    return JsonResponse(
        {
            "status": "healthy",
            "service": "cert-platform",
            "date": timezone.localdate().isoformat(),
        }
    )


@require_GET
def dashboard(request):
    return render(request, "certs/index.html", build_dashboard_context())


@require_GET
def dashboard_api(request):
    context = build_dashboard_context()
    return JsonResponse(
        {
            "employee_count": context["employee_count"],
            "certificate_count": context["certificate_count"],
            "valid_count": context["valid_count"],
            "expiring_count": context["expiring_count"],
            "pending_task_count": context["pending_task_count"],
            "unread_notification_count": context["unread_notification_count"],
        }
    )


@require_GET
def employee_detail(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    certificates = employee.certificates.all().order_by("-expiry_date")
    query_logs = QueryExecution.objects.filter(task__employee=employee).select_related("task")[:10]
    context = {
        "employee": employee,
        "certificates": certificates,
        "query_logs": query_logs,
        "cert_count": certificates.count(),
        "valid_count": certificates.filter(status=Certificate.Status.VALID).count(),
        "expired_count": certificates.filter(status=Certificate.Status.EXPIRED).count(),
    }
    return render(request, "certs/employee_detail.html", context)


@require_GET
def expiring(request):
    days = max(1, min(int(request.GET.get("days", 60)), 365))
    today = timezone.localdate()
    certificates = Certificate.objects.select_related("employee").filter(
        status=Certificate.Status.VALID,
        expiry_date__isnull=False,
        expiry_date__gt=today,
        expiry_date__lte=today + timedelta(days=days),
    ).order_by("expiry_date")
    return render(request, "certs/expiring.html", {"certificates": certificates, "days": days})


@require_GET
def logs(request):
    logs_qs = QueryExecution.objects.select_related("task", "task__employee").order_by("-created_at")[:100]
    return render(request, "certs/logs.html", {"logs": logs_qs})


@require_http_methods(["GET", "POST"])
def search(request):
    search_type = request.POST.get("type", "name") if request.method == "POST" else request.GET.get("type", "name")
    keyword = (request.POST.get("keyword", "") if request.method == "POST" else request.GET.get("keyword", "")).strip()
    employees = None
    suggestions_list = None
    if keyword:
        if search_type == "id_card":
            employees = Employee.objects.filter(id_number__icontains=keyword)[:50]
        else:
            results = list(Employee.objects.filter(name__icontains=keyword)[:50])
            if len(results) > 1:
                suggestions_list = results
            else:
                employees = results
    return render(
        request,
        "certs/search.html",
        {
            "employees": employees,
            "keyword": keyword,
            "search_type": search_type,
            "suggestions_list": suggestions_list,
        },
    )


@require_GET
def employee_add_page(request):
    return render(request, "certs/employee_add.html")


@require_GET
def certificate_add_page(request):
    return render(request, "certs/certificate_add.html", {"employees": Employee.objects.order_by("name")})


@csrf_exempt
@require_http_methods(["POST"])
def api_add_employee(request):
    try:
        payload = _normalize_employee_payload(_request_data(request))
        if Employee.objects.filter(id_number=payload["id_number"]).exists():
            return JsonResponse({"success": False, "error": "该身份证号已存在"}, status=400)
        employee = Employee.objects.create(**payload)
        return JsonResponse({"success": True, "employee_id": employee.id})
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)


@require_GET
def api_employees_list(request):
    employees = [_employee_dict(emp) for emp in Employee.objects.order_by("name")]
    return JsonResponse(employees, safe=False)


@require_GET
def api_employees_search(request):
    keyword = request.GET.get("keyword", "").strip()
    if not keyword:
        return JsonResponse([], safe=False)
    employees = [_employee_dict(emp) for emp in Employee.objects.filter(Q(name__icontains=keyword) | Q(id_number__icontains=keyword))[:10]]
    return JsonResponse(employees, safe=False)


@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def api_employee_detail(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    if request.method == "GET":
        return JsonResponse(_employee_dict(employee))
    if request.method == "DELETE":
        employee.delete()
        return JsonResponse({"success": True})
    try:
        payload = _normalize_employee_payload(_request_data(request))
        if Employee.objects.exclude(id=employee.id).filter(id_number=payload["id_number"]).exists():
            return JsonResponse({"success": False, "error": "该身份证号已存在"}, status=400)
        for key, value in payload.items():
            setattr(employee, key, value)
        employee.save()
        return JsonResponse({"success": True})
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def api_add_certificate(request):
    try:
        payload = _normalize_certificate_payload(_request_data(request))
        employee = Employee.objects.get(id=payload.pop("employee_id"))
        cert = Certificate.objects.create(employee=employee, **payload)
        return JsonResponse({"success": True, "certificate_id": cert.id})
    except Employee.DoesNotExist:
        return JsonResponse({"success": False, "error": "员工不存在"}, status=400)
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)


@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def api_certificate_detail(request, cert_id):
    cert = get_object_or_404(Certificate.objects.select_related("employee"), id=cert_id)
    if request.method == "GET":
        return JsonResponse(_certificate_dict(cert))
    if request.method == "DELETE":
        cert.delete()
        return JsonResponse({"success": True})
    try:
        payload = _normalize_certificate_payload(_request_data(request))
        cert.employee = Employee.objects.get(id=payload.pop("employee_id"))
        for key, value in payload.items():
            setattr(cert, key, value)
        cert.save()
        return JsonResponse({"success": True})
    except Employee.DoesNotExist:
        return JsonResponse({"success": False, "error": "员工不存在"}, status=400)
    except ValueError as exc:
        return JsonResponse({"success": False, "error": str(exc)}, status=400)


@require_GET
def queue(request):
    filter_status = request.GET.get("status", "").strip()
    tasks = QueryTask.objects.select_related("employee")
    if filter_status:
        tasks = tasks.filter(status=filter_status)
    counts = {
        "pending": QueryTask.objects.filter(status=QueryTask.Status.PENDING).count(),
        "processing": QueryTask.objects.filter(status=QueryTask.Status.RUNNING).count(),
        "completed": QueryTask.objects.filter(status=QueryTask.Status.SUCCESS).count(),
        "failed": QueryTask.objects.filter(status=QueryTask.Status.FAILED).count(),
    }
    statistics = list(
        QueryTask.objects.values("created_at__date")
        .order_by("-created_at__date")[:7]
    )
    return render(
        request,
        "certs/queue.html",
        {"tasks": tasks[:200], "counts": counts, "statistics": statistics, "filter_status": filter_status},
    )


@require_POST
def queue_add(request):
    employees_text = request.POST.get("employees", "")
    query_type = request.POST.get("query_type", QueryTask.QueryType.ALL)
    priority = int(request.POST.get("priority", 5))
    added = 0
    duplicate = 0
    not_found = 0
    for line in employees_text.strip().splitlines():
        token = line.strip().split()[0] if line.strip() else ""
        if not token:
            continue
        employee = None
        if token.isdigit():
            employee = Employee.objects.filter(id=int(token)).first()
        if employee is None:
            employee = Employee.objects.filter(name=token).first() or Employee.objects.filter(name__icontains=token).first()
        if employee is None:
            not_found += 1
            continue
        if QueryTask.objects.filter(employee=employee, status__in=[QueryTask.Status.PENDING, QueryTask.Status.RUNNING]).exists():
            duplicate += 1
            continue
        QueryTask.objects.create(employee=employee, query_type=query_type, priority=priority)
        added += 1
    return JsonResponse(
        {
            "success": True,
            "added": added,
            "duplicate": duplicate,
            "not_found": not_found,
            "message": f"添加完成: 成功 {added} 条, 重复 {duplicate} 条, 未找到 {not_found} 条",
        }
    )


@require_POST
def queue_clear(request):
    deleted, _ = QueryTask.objects.filter(status__in=[QueryTask.Status.SUCCESS, QueryTask.Status.FAILED, QueryTask.Status.EMPTY]).delete()
    return JsonResponse({"success": True, "deleted": deleted})


@require_GET
def queue_process(request):
    pending_count = QueryTask.objects.filter(status=QueryTask.Status.PENDING).count()
    if pending_count == 0:
        return redirect("queue")
    return render(request, "certs/queue_process.html", {"pending_count": pending_count})


@require_GET
def queue_stats(request):
    return JsonResponse(
        {
            "pending": QueryTask.objects.filter(status=QueryTask.Status.PENDING).count(),
            "processing": QueryTask.objects.filter(status=QueryTask.Status.RUNNING).count(),
            "completed": QueryTask.objects.filter(status=QueryTask.Status.SUCCESS).count(),
            "failed": QueryTask.objects.filter(status=QueryTask.Status.FAILED).count(),
        }
    )


@require_POST
def api_queue_trigger(request):
    pending_count = QueryTask.objects.filter(status=QueryTask.Status.PENDING).count()
    if pending_count == 0:
        return JsonResponse({"success": False, "error": "没有待处理任务"})
    processed = process_pending_tasks_inline()
    return JsonResponse(
        {
            "success": True,
            "pending_count": pending_count,
            "processed_count": processed,
            "message": f"已处理{processed}条任务",
        }
    )


@require_GET
def employees_page(request):
    search = request.GET.get("search", "").strip()
    employees = Employee.objects.order_by("-id")
    if search:
        employees = employees.filter(Q(name__icontains=search) | Q(id_number__icontains=search) | Q(department__icontains=search))
    return render(request, "certs/employees.html", {"employees": employees[:100], "search": search})


def employees_dispatch(request):
    if request.method == "GET":
        return employees_page(request)
    if request.method == "POST":
        return add_employee(request)
    return HttpResponse(status=405)


@require_POST
def add_employee(request):
    try:
        payload = _normalize_employee_payload(request.POST.dict())
        if Employee.objects.filter(id_number=payload["id_number"]).exists():
            raise ValueError("该身份证号已存在")
        Employee.objects.create(**payload)
        return redirect("employees_page")
    except ValueError as exc:
        return HttpResponse(f'<script>alert("添加失败: {exc}");history.back();</script>')


@require_POST
def edit_employee(request, emp_id):
    employee = get_object_or_404(Employee, id=emp_id)
    try:
        payload = _normalize_employee_payload(request.POST.dict())
        if Employee.objects.exclude(id=employee.id).filter(id_number=payload["id_number"]).exists():
            raise ValueError("该身份证号已存在")
        for key, value in payload.items():
            setattr(employee, key, value)
        employee.save()
        return redirect("employees_page")
    except ValueError as exc:
        return HttpResponse(f'<script>alert("更新失败: {exc}");history.back();</script>')


@require_POST
def delete_employee(request, emp_id):
    get_object_or_404(Employee, id=emp_id).delete()
    return redirect("employees_page")


@require_GET
def certificates_page(request):
    employee_id = request.GET.get("employee_id")
    search = request.GET.get("search", "").strip()
    certificates = Certificate.objects.select_related("employee").order_by("-id")
    if employee_id:
        certificates = certificates.filter(employee_id=employee_id)
    if search:
        certificates = certificates.filter(Q(certificate_type__icontains=search) | Q(operation_item__icontains=search))
    return render(
        request,
        "certs/certificates.html",
        {
            "certificates": certificates[:200],
            "employees": Employee.objects.order_by("name"),
            "selected_employee_id": int(employee_id) if employee_id else None,
            "search": search,
        },
    )


def certificates_dispatch(request):
    if request.method == "GET":
        return certificates_page(request)
    if request.method == "POST":
        return add_certificate(request)
    return HttpResponse(status=405)


@require_POST
def add_certificate(request):
    try:
        payload = _normalize_certificate_payload(request.POST.dict())
        employee = Employee.objects.get(id=payload.pop("employee_id"))
        Certificate.objects.create(employee=employee, **payload)
        return redirect("certificates_page")
    except (ValueError, Employee.DoesNotExist) as exc:
        return HttpResponse(f'<script>alert("添加失败: {exc}");history.back();</script>')


@require_POST
def edit_certificate(request, cert_id):
    cert = get_object_or_404(Certificate, id=cert_id)
    try:
        payload = _normalize_certificate_payload(request.POST.dict())
        cert.employee = Employee.objects.get(id=payload.pop("employee_id"))
        for key, value in payload.items():
            setattr(cert, key, value)
        cert.save()
        return redirect("certificates_page")
    except (ValueError, Employee.DoesNotExist) as exc:
        return HttpResponse(f'<script>alert("更新失败: {exc}");history.back();</script>')


@require_POST
def delete_certificate(request, cert_id):
    get_object_or_404(Certificate, id=cert_id).delete()
    return redirect("certificates_page")


@require_GET
def api_notifications(request):
    data = [
        {
            "id": n.id,
            "title": n.title,
            "content": n.content,
            "notification_type": n.notification_type,
            "created_at": n.created_at.isoformat(),
        }
        for n in Notification.objects.filter(is_read=False).order_by("-created_at")[:10]
    ]
    return JsonResponse({"notifications": data})


@require_POST
def api_notifications_read(request):
    updated = Notification.objects.filter(is_read=False).update(is_read=True)
    return JsonResponse({"success": True, "updated": updated})


_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
_HEADER_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)


def _build_excel(headers, rows, sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _BORDER
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical='center')
    return wb


@require_GET
def export_employees(request):
    employees = Employee.objects.order_by("-id")
    wb = _build_excel(
        ['ID', '姓名', '身份证号', '部门', '电话', '岗位', '在职状态', '创建日期'],
        [
            [e.id, e.name, e.id_number, e.department, e.phone, e.position, e.employment_status, e.created_at.strftime("%Y-%m-%d")]
            for e in employees
        ],
        '员工列表',
    )
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=employees_{timezone.now().strftime("%Y%m%d")}.xlsx'
    return response


@require_GET
def export_certificates(request):
    certificates = Certificate.objects.select_related("employee").order_by("-id")
    wb = _build_excel(
        ['ID', '员工姓名', '身份证号', '证书类型', '作业项目', '发证机关', '发证日期', '有效期至', '复审日期', '状态'],
        [
            [
                c.id,
                c.employee.name,
                c.employee.id_number,
                c.certificate_type,
                c.operation_item,
                c.issuing_authority,
                c.issue_date.isoformat() if c.issue_date else "",
                c.expiry_date.isoformat() if c.expiry_date else "",
                c.review_date.isoformat() if c.review_date else "",
                c.status,
            ]
            for c in certificates
        ],
        '证书列表',
    )
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=certificates_{timezone.now().strftime("%Y%m%d")}.xlsx'
    return response

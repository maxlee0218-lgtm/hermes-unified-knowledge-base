#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
证书管理系统 v3.0 - 深度重构版

改进点:
- SQLite 本地数据库, WAL 模式, busy_timeout
- 环境变量配置 + .env 文件支持
- HTTP Basic Auth 认证
- 日期字段严格校验 (validate_date)
- 移除 requests 依赖 (notify_change)
- 自动状态刷新返回计数
- 404 错误返回模板
- 删除操作级联 (ON DELETE CASCADE 已由 DB 保证)
- 导出 Excel 带表头样式
- 健康检查返回版本号和数据库大小
"""

import os
import logging
import secrets
from datetime import datetime, date
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect,
    url_for, jsonify, send_file, abort
)

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import DB_PATH, DATA_DIR, db_connection, get_db, init_db
from services.certificate_service import (
    create_certificate as create_certificate_record,
    delete_certificate as delete_certificate_record,
    get_certificate as get_certificate_record,
    list_certificates_for_page,
    update_certificate as update_certificate_record,
)
from services.dashboard_service import (
    get_dashboard_data,
    get_employee_detail_data,
    get_expiring_data,
    get_recent_logs,
    get_unread_notifications,
    mark_all_notifications_read,
)
from services.employee_service import (
    create_employee as create_employee_record,
    create_employee_from_form,
    delete_employee as delete_employee_record,
    get_employee as get_employee_record,
    list_basic_employees,
    list_employees_for_page,
    search_employees,
    search_page_results,
    update_employee as update_employee_record,
    update_employee_from_form,
)
from services.errors import NotFoundError, ValidationError
from services.query_queue_service import (
    add_queue_tasks,
    clear_finished_queue_tasks,
    get_pending_queue_count,
    get_queue_page_data,
    get_queue_stats,
)
from status_rules import (
    calculate_days_until_expiry,
    get_status_class,
    validate_date,
)

# ============================================================
# 配置
# ============================================================
app = Flask(__name__)

app.secret_key = os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32))

AUTH_USERNAME = os.environ.get('CERT_AUTH_USER', 'admin')
AUTH_PASSWORD = os.environ.get('CERT_AUTH_PASS', 'admin123')

os.makedirs(DATA_DIR, exist_ok=True)

# ============================================================
# 日志
# ============================================================
log_handlers = [logging.StreamHandler()]
LOG_FILE = os.environ.get('CERT_LOG_FILE', '/var/log/cert-system/app.log')
if os.path.isdir(os.path.dirname(LOG_FILE)):
    log_handlers.append(logging.FileHandler(LOG_FILE, encoding='utf-8'))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=log_handlers
)
logger = logging.getLogger('cert_system')

# ============================================================
# 认证
# ============================================================
def check_auth(username, password):
    return username == AUTH_USERNAME and password == AUTH_PASSWORD


@app.before_request
def require_auth():
    if request.path == '/health':
        return None
    auth = request.authorization
    if not auth or not check_auth(auth.username, auth.password):
        return (
            '<h1>需要认证</h1><p>请输入用户名和密码访问系统。</p>',
            401,
            {'WWW-Authenticate': 'Basic realm="Certificate System"'}
        )


# ============================================================
# 工具函数
# ============================================================
def _truncate_date(val):
    """安全截断日期到 10 位"""
    if val:
        s = str(val)
        if len(s) >= 10:
            return s[:10]
    return val or ''


def _parse_form_data():
    """统一解析请求体为 dict"""
    if request.is_json:
        return request.get_json() or {}
    elif request.form:
        return dict(request.form)
    return {}


# ============================================================
# 健康检查
# ============================================================
@app.route('/health')
def health():
    try:
        conn = get_db()
        conn.execute("SELECT 1")
        db_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
        conn.close()
        return jsonify({
            "status": "healthy", "db": "ok",
            "db_size_mb": round(db_size / 1024 / 1024, 2),
            "version": "3.0"
        })
    except Exception as e:
        return jsonify({"status": "unhealthy", "error": str(e)}), 500


# ============================================================
# 首页
# ============================================================
@app.route('/')
def index():
    status_list = request.args.getlist('status')
    cert_type_filter = request.args.get('cert_type', '').strip()
    keyword = request.args.get('keyword', '').strip()
    show_all = request.args.get('show_all', '') == '1'
    dashboard = get_dashboard_data(
        status_list=status_list,
        cert_type_filter=cert_type_filter,
        keyword=keyword,
        show_all=show_all,
    )

    return render_template('index.html',
        certificates=dashboard['certificates'],
        employee_count=dashboard['employee_count'],
        certificate_count=dashboard['certificate_count'],
        valid_count=dashboard['valid_count'],
        expired_count=dashboard['expired_count'],
        warning_count=dashboard['warning_count'],
        status_list=status_list, show_all=show_all,
        cert_type_filter=cert_type_filter, keyword=keyword,
        auto_updated=dashboard['auto_updated']
    )


# ============================================================
# 通知 API
# ============================================================
@app.route('/api/notifications')
def api_notifications():
    return jsonify({'notifications': get_unread_notifications()})


@app.route('/api/notifications/read', methods=['POST'])
def api_notifications_read():
    updated = mark_all_notifications_read()
    return jsonify({'success': True, 'updated': updated})


# ============================================================
# 员工详情
# ============================================================
@app.route('/employee/<int:employee_id>')
def employee_detail(employee_id):
    try:
        detail = get_employee_detail_data(employee_id)
    except NotFoundError:
        abort(404)

    return render_template(
        'employee_detail.html',
        employee=detail['employee'],
        certificates=detail['certificates'],
        query_logs=detail['query_logs'],
        cert_count=detail['cert_count'],
        valid_count=detail['valid_count'],
        expired_count=detail['expired_count'],
    )


# ============================================================
# 即将到期
# ============================================================
@app.route('/expiring')
def expiring():
    data = get_expiring_data(request.args.get('days', 60, type=int))
    return render_template('expiring.html', certificates=data['certificates'], days=data['days'])


# ============================================================
# 查询日志
# ============================================================
@app.route('/logs')
def logs():
    return render_template('logs.html', logs=get_recent_logs())


# ============================================================
# 搜索
# ============================================================
@app.route('/search', methods=['GET', 'POST'])
def search():
    if request.method == 'POST':
        search_type = request.form.get('type', 'name')
        keyword = request.form.get('keyword', '').strip()
    else:
        search_type = request.args.get('type', 'name')
        keyword = request.args.get('keyword', '').strip()

    employees, suggestions_list = search_page_results(search_type, keyword)

    return render_template('search.html',
        employees=employees, keyword=keyword,
        search_type=search_type, suggestions_list=suggestions_list
    )


# ============================================================
# 录入页面
# ============================================================
@app.route('/employees/add')
def employee_add_page():
    return render_template('employee_add.html')


@app.route('/certificates/add')
def certificate_add_page():
    employees = list_basic_employees()
    return render_template('certificate_add.html', employees=employees)


# ============================================================
# 员工 API
# ============================================================
@app.route('/api/employees', methods=['POST'])
def api_add_employee():
    data = _parse_form_data()

    try:
        employee_id = create_employee_record(data)
        logger.info(f"添加员工: {data.get('name', '').strip()} (ID: {employee_id})")
        return jsonify({'success': True, 'employee_id': employee_id})
    except ValidationError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        logger.error(f"添加员工失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/employees/list')
def api_employees_list():
    return jsonify(list_basic_employees())


@app.route('/api/employees/search')
def api_employees_search():
    keyword = request.args.get('keyword', '').strip()
    return jsonify(search_employees(keyword))


@app.route('/api/employees/<int:emp_id>', methods=['GET'])
def api_get_employee(emp_id):
    try:
        return jsonify(get_employee_record(emp_id))
    except NotFoundError as e:
        return jsonify({'error': str(e)}), 404


@app.route('/api/employees/<int:emp_id>', methods=['PUT'])
def api_update_employee(emp_id):
    data = _parse_form_data()

    try:
        update_employee_record(emp_id, data)
        logger.info(f"更新员工: ID={emp_id}")
        return jsonify({'success': True})
    except ValidationError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        logger.error(f"更新员工失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/employees/<int:emp_id>', methods=['DELETE'])
def api_delete_employee(emp_id):
    try:
        delete_employee_record(emp_id)
        logger.info(f"删除员工: ID={emp_id}")
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"删除员工失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# 证书 API
# ============================================================
@app.route('/api/certificates', methods=['POST'])
def api_add_certificate():
    data = _parse_form_data()

    try:
        certificate_id = create_certificate_record(data, include_review_date=True)
        logger.info(
            f"添加证书: type={data.get('certificate_type', '').strip()}, employee_id={data.get('employee_id')}"
        )
        return jsonify({'success': True, 'certificate_id': certificate_id})
    except ValidationError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        logger.error(f"添加证书失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/certificates/<int:cert_id>', methods=['GET'])
def api_get_certificate(cert_id):
    try:
        return jsonify(get_certificate_record(cert_id))
    except NotFoundError as e:
        return jsonify({'error': str(e)}), 404


@app.route('/api/certificates/<int:cert_id>', methods=['PUT'])
def api_update_certificate(cert_id):
    data = _parse_form_data()

    try:
        update_certificate_record(cert_id, data)
        logger.info(f"更新证书: ID={cert_id}")
        return jsonify({'success': True})
    except ValidationError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        logger.error(f"更新证书失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/certificates/<int:cert_id>', methods=['DELETE'])
def api_delete_certificate(cert_id):
    try:
        delete_certificate_record(cert_id)
        logger.info(f"删除证书: ID={cert_id}")
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"删除证书失败: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# 查询队列
# ============================================================
@app.route('/queue')
def queue():
    data = get_queue_page_data(request.args.get('status', '').strip())
    return render_template(
        'queue.html',
        tasks=data['tasks'],
        counts=data['counts'],
        statistics=data['statistics'],
        filter_status=data['filter_status'],
    )


@app.route('/queue/add', methods=['POST'])
def queue_add():
    employees_text = request.form.get('employees', '')
    query_type = request.form.get('query_type', '全部')
    priority = request.form.get('priority', 5, type=int)
    result = add_queue_tasks(employees_text, query_type=query_type, priority=priority)
    logger.info(
        f"队列添加: 成功={result['added']}, 重复={result['duplicate']}, 未找到={result['not_found']}"
    )
    return jsonify({
        'success': True,
        'added': result['added'],
        'duplicate': result['duplicate'],
        'not_found': result['not_found'],
        'message': (
            f"添加完成: 成功 {result['added']} 条, 重复 {result['duplicate']} 条, 未找到 {result['not_found']} 条"
        )
    })


@app.route('/queue/clear', methods=['POST'])
def queue_clear():
    deleted = clear_finished_queue_tasks()
    logger.info(f"队列清理: 删除 {deleted} 条")
    return jsonify({'success': True, 'deleted': deleted})


@app.route('/queue/process')
def queue_process():
    pending_count = get_pending_queue_count()

    if pending_count == 0:
        return redirect(url_for('queue'))

    return render_template('queue_process.html', pending_count=pending_count)


@app.route('/api/queue/stats')
def queue_stats():
    return jsonify(get_queue_stats())


@app.route('/api/queue/trigger', methods=['POST'])
def api_queue_trigger():
    pending_count = get_pending_queue_count()

    if pending_count == 0:
        return jsonify({'success': False, 'error': '没有待处理任务'})

    marker_file = '/tmp/trigger_queue_process'
    try:
        with open(marker_file, 'w') as f:
            f.write(str(pending_count))
    except OSError:
        pass

    return jsonify({
        'success': True,
        'pending_count': pending_count,
        'message': f'已触发处理{pending_count}条任务'
    })


# ============================================================
# 员工管理（页面 + 表单）
# ============================================================
@app.route('/employees')
def employees_page():
    search = request.args.get('search', '').strip()
    employees = list_employees_for_page(search=search)
    for emp in employees:
        emp['created_at'] = _truncate_date(emp.get('created_at'))
        emp['updated_at'] = _truncate_date(emp.get('updated_at'))

    return render_template('employees.html', employees=employees, search=search)


@app.route('/employees', methods=['POST'])
def add_employee():
    name = request.form.get('name', '').strip()
    id_card = request.form.get('id_card', '').strip()
    department = request.form.get('department', '').strip()
    phone = request.form.get('phone', '').strip()

    if not name or not id_card:
        return '<script>alert("姓名和身份证号为必填项");history.back();</script>'

    try:
        create_employee_from_form(name, id_card, department, phone)
        logger.info(f"添加员工(表单): {name}")
        return redirect(url_for('employees_page'))
    except ValidationError as e:
        return f'<script>alert("添加失败: {e}");history.back();</script>'
    except Exception as e:
        logger.error(f"添加员工失败: {e}")
        return f'<script>alert("添加失败: {e}");history.back();</script>'


@app.route('/employees/<int:emp_id>/edit', methods=['POST'])
def edit_employee(emp_id):
    name = request.form.get('name', '').strip()
    id_card = request.form.get('id_card', '').strip()
    department = request.form.get('department', '').strip()
    phone = request.form.get('phone', '').strip()

    if not name or not id_card:
        return '<script>alert("姓名和身份证号为必填项");history.back();</script>'

    try:
        update_employee_from_form(emp_id, name, id_card, department, phone)
        logger.info(f"编辑员工: ID={emp_id}")
        return redirect(url_for('employees_page'))
    except ValidationError as e:
        return f'<script>alert("更新失败: {e}");history.back();</script>'
    except Exception as e:
        logger.error(f"编辑员工失败: {e}")
        return f'<script>alert("更新失败: {e}");history.back();</script>'


@app.route('/employees/<int:emp_id>/delete', methods=['POST'])
def delete_employee(emp_id):
    try:
        delete_employee_record(emp_id)
        logger.info(f"删除员工: ID={emp_id}")
        return redirect(url_for('employees_page'))
    except Exception as e:
        logger.error(f"删除员工失败: {e}")
        return f'<script>alert("删除失败: {e}");history.back();</script>'


# ============================================================
# 证书管理（页面 + 表单）
# ============================================================
@app.route('/certificates')
def certificates_page():
    employee_id = request.args.get('employee_id', type=int)
    search = request.args.get('search', '').strip()
    certificates = list_certificates_for_page(employee_id=employee_id, search=search)
    for cert in certificates:
        for f in ('expiry_date', 'issue_date', 'created_at'):
            cert[f] = _truncate_date(cert.get(f))
    employees = list_basic_employees()

    return render_template('certificates.html',
        certificates=certificates, employees=employees,
        selected_employee_id=employee_id, search=search
    )


@app.route('/certificates', methods=['POST'])
def add_certificate():
    employee_id = request.form.get('employee_id', type=int)
    certificate_type = request.form.get('certificate_type', '').strip()

    if not employee_id or not certificate_type:
        return '<script>alert("员工和证书类型为必填项");history.back();</script>'

    try:
        create_certificate_record(request.form, include_review_date=False)
        logger.info(f"添加证书(表单): type={certificate_type}")
        return redirect(url_for('certificates_page'))
    except ValidationError as e:
        return f'<script>alert("添加失败: {e}");history.back();</script>'
    except Exception as e:
        logger.error(f"添加证书失败: {e}")
        return f'<script>alert("添加失败: {e}");history.back();</script>'


@app.route('/certificates/<int:cert_id>/edit', methods=['POST'])
def edit_certificate(cert_id):
    employee_id = request.form.get('employee_id', type=int)
    certificate_type = request.form.get('certificate_type', '').strip()

    if not employee_id or not certificate_type:
        return '<script>alert("员工和证书类型为必填项");history.back();</script>'

    try:
        update_certificate_record(cert_id, request.form)
        logger.info(f"编辑证书: ID={cert_id}")
        return redirect(url_for('certificates_page'))
    except ValidationError as e:
        return f'<script>alert("更新失败: {e}");history.back();</script>'
    except Exception as e:
        logger.error(f"编辑证书失败: {e}")
        return f'<script>alert("更新失败: {e}");history.back();</script>'


@app.route('/certificates/<int:cert_id>/delete', methods=['POST'])
def delete_certificate(cert_id):
    try:
        delete_certificate_record(cert_id)
        logger.info(f"删除证书: ID={cert_id}")
        return redirect(url_for('certificates_page'))
    except Exception as e:
        logger.error(f"删除证书失败: {e}")
        return f'<script>alert("删除失败: {e}");history.back();</script>'


# ============================================================
# 导出功能
# ============================================================
_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
_HEADER_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)


def _build_excel(headers, rows, sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # 表头样式
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _BORDER

    # 数据行
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical='center')

    _auto_width(ws)
    return wb


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        ws.column_dimensions[column].width = min((max_length + 4), 40)


@app.route('/export/employees')
def export_employees():
    with db_connection() as conn:
        employees = conn.execute('SELECT * FROM employees ORDER BY id DESC').fetchall()

    headers = ['ID', '姓名', '身份证号', '部门', '电话', '岗位', '在职状态', '创建日期']
    rows = []
    for emp in employees:
        e = dict(emp)
        rows.append([
            e.get('id', ''), e.get('name', ''), e.get('id_number', ''),
            e.get('department', ''), e.get('phone', ''), e.get('position', ''),
            e.get('employment_status', ''),
            _truncate_date(e.get('created_at'))
        ])

    wb = _build_excel(headers, rows, '员工列表')
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("导出员工列表")
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'employees_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/export/certificates')
def export_certificates():
    with db_connection() as conn:
        certificates = conn.execute('''
            SELECT c.*, e.name as employee_name, e.id_number
            FROM certificates c JOIN employees e ON c.employee_id = e.id ORDER BY c.id DESC
        ''').fetchall()

    headers = ['ID', '员工姓名', '身份证号', '证书类型', '作业项目', '发证机关',
               '发证日期', '有效期至', '复审日期', '状态']
    rows = []
    for c in certificates:
        d = dict(c)
        rows.append([
            d.get('id', ''), d.get('employee_name', ''), d.get('id_number', ''),
            d.get('certificate_type', ''), d.get('operation_item', ''),
            d.get('issuing_authority', ''),
            _truncate_date(d.get('issue_date')), _truncate_date(d.get('expiry_date')),
            _truncate_date(d.get('review_date')), d.get('status', '')
        ])

    wb = _build_excel(headers, rows, '证书列表')
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("导出证书列表")
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'certificates_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ============================================================
# 错误处理
# ============================================================
@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(500)
def server_error(e):
    logger.error(f"服务器错误: {e}", exc_info=True)
    return render_template('500.html'), 500


# ============================================================
# 入口
# ============================================================
if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
else:
    init_db()

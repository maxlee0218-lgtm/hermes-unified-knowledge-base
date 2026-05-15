from flask import Blueprint, send_file, jsonify
from models import Employee, Certificate
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

export_bp = Blueprint('export', __name__)

def create_excel_response(wb, filename):
    """创建Excel响应"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@export_bp.route('/employees', methods=['GET'])
def export_employees():
    """导出员工列表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '员工列表'
    
    # 设置表头
    headers = ['ID', '姓名', '身份证号', '部门', '职位', '证书数量', '创建时间']
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 获取数据
    employees = Employee.query.all()
    
    for emp in employees:
        ws.append([
            emp.id,
            emp.name,
            emp.id_number,
            emp.department or '',
            emp.position or '',
            len(emp.certificates),
            emp.created_at.strftime('%Y-%m-%d %H:%M:%S') if emp.created_at else ''
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    
    filename = f'员工列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(wb, filename)

@export_bp.route('/certificates', methods=['GET'])
def export_certificates():
    """导出证书列表"""
    wb = Workbook()
    ws = wb.active
    ws.title = '证书列表'
    
    # 设置表头
    headers = ['ID', '员工姓名', '身份证号', '证书类型', '作业类别', '操作项目', 
               '证书编号', '发证机关', '发证日期', '有效期起', '有效期止', '复审日期', '状态']
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 获取数据
    certs = Certificate.query.all()
    
    for cert in certs:
        ws.append([
            cert.id,
            cert.employee.name if cert.employee else '',
            cert.employee.id_number if cert.employee else '',
            cert.cert_type,
            cert.cert_category or '',
            cert.cert_item or '',
            cert.cert_number or '',
            cert.issue_authority or '',
            cert.issue_date.strftime('%Y-%m-%d') if cert.issue_date else '',
            cert.valid_from.strftime('%Y-%m-%d') if cert.valid_from else '',
            cert.valid_until.strftime('%Y-%m-%d') if cert.valid_until else '',
            cert.review_date.strftime('%Y-%m-%d') if cert.review_date else '',
            cert.status
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 10
    
    filename = f'证书列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(wb, filename)

@export_bp.route('/expiring', methods=['GET'])
def export_expiring():
    """导出即将到期证书报表"""
    from flask import request
    from datetime import timedelta
    
    days = request.args.get('days', 30, type=int)
    today = datetime.now().date()
    deadline = today + timedelta(days=days)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f'即将到期证书({days}天)'
    
    # 设置表头
    headers = ['员工姓名', '身份证号', '部门', '证书类型', '作业类别', '操作项目', 
               '证书编号', '有效期止', '剩余天数', '状态']
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 获取数据
    certs = Certificate.query.filter(
        Certificate.valid_until <= deadline,
        Certificate.valid_until >= today
    ).order_by(Certificate.valid_until).all()
    
    for cert in certs:
        remaining_days = (cert.valid_until - today).days if cert.valid_until else 0
        ws.append([
            cert.employee.name if cert.employee else '',
            cert.employee.id_number if cert.employee else '',
            cert.employee.department if cert.employee else '',
            cert.cert_type,
            cert.cert_category or '',
            cert.cert_item or '',
            cert.cert_number or '',
            cert.valid_until.strftime('%Y-%m-%d') if cert.valid_until else '',
            remaining_days,
            cert.status
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10
    
    filename = f'即将到期证书报表_{days}天_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return create_excel_response(wb, filename)

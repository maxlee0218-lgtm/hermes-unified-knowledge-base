#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import logging
import os
import secrets
import json
from base64 import b64decode
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import Flask, abort, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import case, func, or_, text
from sqlalchemy.exc import OperationalError

from config import Config
from models import (
    Certificate,
    Employee,
    QueryLog,
    QueryQueue,
    SystemConfig,
    UpdateNotification,
    db,
    ensure_schema,
    format_date,
    format_datetime,
)
from services.query_service import QueryService, evaluate_certificate_status, normalize_query_type

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("cert_system_operational")

BASE_DIR = os.path.abspath(os.path.dirname(__file__))


def create_app(uri_override=None, fallback_active=False):
    app = Flask(__name__, template_folder="templates", static_folder="static")
    app.config.from_object(Config)
    app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))
    base_uri = uri_override or os.environ.get(
        "SQLALCHEMY_DATABASE_URI", app.config["SQLALCHEMY_DATABASE_URI"]
    )
    app.config["SQLALCHEMY_DATABASE_URI"] = base_uri
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    if app.config["SQLALCHEMY_DATABASE_URI"].startswith("mysql"):
        app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
            "pool_pre_ping": True,
            "connect_args": {
                "connect_timeout": 5,
                "read_timeout": 15,
                "write_timeout": 15,
            },
        }

    db.init_app(app)
    app.config["ACTIVE_DATABASE_URI"] = app.config["SQLALCHEMY_DATABASE_URI"]
    app.config["DATABASE_FALLBACK_ACTIVE"] = fallback_active
    try:
        ensure_schema(app)
    except OperationalError as exc:
        if app.config["SQLALCHEMY_DATABASE_URI"].startswith("mysql") and not fallback_active:
            fallback_path = os.environ.get(
                "CERT_SQLITE_FALLBACK_PATH",
                os.path.join(BASE_DIR, "data", "certificate_fallback.db"),
            )
            fallback_uri = f"sqlite:///{fallback_path}"
            logger.warning(
                "MariaDB 不可达，切换到本地 SQLite 兜底: %s; error=%s",
                fallback_uri,
                exc,
            )
            return create_app(uri_override=fallback_uri, fallback_active=True)
        else:
            raise

    auth_user = os.environ.get("CERT_AUTH_USER", "admin")
    auth_pass = os.environ.get("CERT_AUTH_PASS", "Cert@2026!Secure")

    @app.before_request
    def require_auth():
        if request.path in {"/health"} or request.path.startswith("/static/"):
            return None
        auth = request.authorization
        if not auth and request.headers.get("Authorization", "").startswith("Basic "):
            try:
                raw = request.headers.get("Authorization", "").split(" ", 1)[1]
                user_pass = b64decode(raw).decode("utf-8")
                username, password = user_pass.split(":", 1)
                auth = type("Auth", (), {"username": username, "password": password})()
            except Exception:
                auth = None
        if not auth or auth.username != auth_user or auth.password != auth_pass:
            return (
                "<h1>需要认证</h1><p>请输入用户名和密码访问系统。</p>",
                401,
                {"WWW-Authenticate": 'Basic realm="Certificate System"'},
            )
        return None

    @app.context_processor
    def inject_now():
        return {"now": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

    register_routes(app)
    return app


def register_routes(app):
    @app.route("/health")
    def health():
        try:
            db.session.execute(text("SELECT 1"))
            return jsonify(
                {
                    "status": "healthy",
                    "service": "cert-platform",
                    "date": date.today().isoformat(),
                    "database": "ok",
                    "database_uri": app.config.get("ACTIVE_DATABASE_URI"),
                    "database_fallback": app.config.get("DATABASE_FALLBACK_ACTIVE", False),
                }
            )
        except Exception as exc:
            return jsonify({"status": "unhealthy", "error": str(exc)}), 500

    @app.route("/")
    def index():
        status_list = request.args.getlist("status")
        cert_type_filter = request.args.get("cert_type", "").strip()
        keyword = request.args.get("keyword", "").strip()
        show_all = request.args.get("show_all", "") == "1"

        employees = Employee.query
        if keyword:
            employees = employees.filter(
                or_(Employee.name.contains(keyword), Employee.id_number.contains(keyword))
            )
        employee_map = {emp.id: emp for emp in employees.order_by(Employee.name.asc()).all()}

        cert_query = Certificate.query.filter(Certificate.employee_id.in_(employee_map.keys() or [0]))
        if cert_type_filter:
            cert_query = cert_query.filter(Certificate.certificate_type == cert_type_filter)

        certificates = []
        grouped = {}
        for cert in cert_query.order_by(Certificate.employee_id.asc(), Certificate.valid_until.desc()).all():
            cert.sync_alias_fields()
            cert.status = evaluate_certificate_status(cert)
            key = (cert.employee_id, cert.certificate_type)
            grouped.setdefault(key, []).append(cert)
        db.session.commit()

        for key, certs in grouped.items():
            if show_all:
                chosen = certs
            else:
                chosen = [certs[0]]
            for cert in chosen:
                row = certificate_row(cert)
                if status_list and row["display_status"] not in status_list:
                    continue
                certificates.append(row)

        # Preserve employees with no certs in search results.
        employee_ids_with_certs = {row["employee_id"] for row in certificates}
        for emp in employee_map.values():
            if emp.id not in employee_ids_with_certs and not cert_type_filter and not status_list:
                certificates.append(
                    {
                        "employee_id": emp.id,
                        "certificate_id": None,
                        "name": emp.name,
                        "id_card": emp.id_number,
                        "certificate_type": "",
                        "operation_item": "",
                        "issuing_authority": "",
                        "review_date": None,
                        "actual_review_date": None,
                        "expiry_date": None,
                        "days_until_expiry": None,
                        "days_to_review": None,
                        "display_status": "",
                        "status_class": "",
                    }
                )

        all_certs = Certificate.query.all()
        valid_count = 0
        expired_count = 0
        warning_count = 0
        for cert in all_certs:
            status = evaluate_certificate_status(cert)
            if status == "有效":
                valid_count += 1
            elif status == "已过期":
                expired_count += 1
            elif status in {"即将到期", "待复审"}:
                warning_count += 1

        return render_template(
            "index.html",
            certificates=sorted(certificates, key=lambda x: (x["name"], x["certificate_type"], x["expiry_date"] or "")),
            employee_count=len(employee_map),
            certificate_count=len(all_certs),
            valid_count=valid_count,
            expired_count=expired_count,
            warning_count=warning_count,
            status_list=status_list,
            cert_type_filter=cert_type_filter,
            keyword=keyword,
            show_all=show_all,
            auto_updated=len(all_certs),
        )

    @app.route("/api/dashboard/")
    def dashboard_api():
        queue_counts = queue_counts_map()
        return jsonify(
            {
                "employee_count": Employee.query.count(),
                "cert_count": Certificate.query.count(),
                "queue_stats": queue_counts,
                "expiring_count": expiring_query(60).count(),
            }
        )

    @app.route("/api/notifications")
    def api_notifications():
        notifications = (
            UpdateNotification.query.filter_by(is_read=False)
            .order_by(UpdateNotification.created_at.desc())
            .limit(10)
            .all()
        )
        return jsonify({"notifications": [notification_to_dict(item) for item in notifications]})

    @app.route("/api/notifications/read", methods=["POST"])
    def api_notifications_read():
        updated = UpdateNotification.query.filter_by(is_read=False).update({"is_read": True})
        db.session.commit()
        return jsonify({"success": True, "updated": updated})

    @app.route("/employees")
    def employees_page():
        search = request.args.get("search", "").strip()
        query = Employee.query
        if search:
            query = query.filter(
                or_(
                    Employee.name.contains(search),
                    Employee.id_number.contains(search),
                    Employee.department.contains(search),
                    Employee.employee_code.contains(search),
                )
            )
        employees = [emp.to_dict() for emp in query.order_by(Employee.created_at.desc()).all()]
        return render_template("employees.html", employees=employees, search=search)

    @app.route("/employee/<int:employee_id>")
    def employee_detail(employee_id):
        employee = Employee.query.get_or_404(employee_id)
        certs = [detail_certificate_row(cert) for cert in Certificate.query.filter_by(employee_id=employee_id).order_by(Certificate.valid_until.desc()).all()]
        logs = [log.to_dict() for log in QueryLog.query.filter_by(employee_id=employee_id).order_by(QueryLog.queried_at.desc()).limit(10).all()]
        emp = employee.to_dict()
        return render_template(
            "employee_detail.html",
            employee=emp,
            certificates=certs,
            query_logs=logs,
            cert_count=len(certs),
            valid_count=sum(1 for c in certs if c["status"] == "有效"),
            expired_count=sum(1 for c in certs if c["status"] == "已过期"),
        )

    @app.route("/search", methods=["GET", "POST"])
    def search():
        if request.method == "POST":
            search_type = request.form.get("type", "name")
            keyword = request.form.get("keyword", "").strip()
        else:
            search_type = request.args.get("type", "name")
            keyword = request.args.get("keyword", "").strip()

        employees = None
        suggestions_list = []
        if keyword:
            if search_type == "id_card":
                employees = [
                    employee_search_row(emp)
                    for emp in Employee.query.filter(Employee.id_number.contains(keyword))
                    .order_by(Employee.name.asc())
                    .all()
                ]
            else:
                matched = (
                    Employee.query.filter(Employee.name.contains(keyword))
                    .order_by(Employee.name.asc())
                    .all()
                )
                suggestions_list = [employee_search_row(emp) for emp in matched[:20]]
                employees = [employee_search_row(emp) for emp in matched]
        return render_template(
            "search.html",
            employees=employees,
            keyword=keyword,
            search_type=search_type,
            suggestions_list=suggestions_list,
        )

    @app.route("/expiring")
    def expiring():
        days = request.args.get("days", 60, type=int)
        certs = [expiring_row(cert) for cert in expiring_query(days).all()]
        return render_template("expiring.html", certificates=certs, days=days)

    @app.route("/logs")
    def logs():
        logs = [log.to_dict() for log in QueryLog.query.order_by(QueryLog.queried_at.desc()).limit(100).all()]
        return render_template("logs.html", logs=logs)

    @app.route("/queue")
    def queue():
        filter_status = request.args.get("status", "").strip()
        query = QueryQueue.query
        if filter_status:
            query = query.filter(QueryQueue.status == filter_status)
        tasks = [task.to_dict() for task in query.order_by(QueryQueue.created_at.desc()).all()]

        statistics = (
            db.session.query(
                func.date(QueryQueue.created_at).label("query_date"),
                func.count(QueryQueue.id).label("total_tasks"),
                func.sum(case((QueryQueue.status == "已完成", 1), else_=0)).label("completed"),
                func.sum(case((QueryQueue.status == "失败", 1), else_=0)).label("failed"),
                func.sum(case((QueryQueue.status == "无记录", 1), else_=0)).label("no_record"),
                func.sum(QueryQueue.new_cert_count).label("total_new_certs"),
                func.avg(QueryQueue.duration_seconds).label("avg_duration"),
            )
            .group_by(func.date(QueryQueue.created_at))
            .order_by(func.date(QueryQueue.created_at).desc())
            .limit(14)
            .all()
        )
        statistics_rows = [
            {
                "query_date": str(item.query_date),
                "total_tasks": int(item.total_tasks or 0),
                "completed": int(item.completed or 0),
                "failed": int(item.failed or 0),
                "no_record": int(item.no_record or 0),
                "total_new_certs": int(item.total_new_certs or 0),
                "avg_duration": float(item.avg_duration or 0),
            }
            for item in statistics
        ]
        return render_template(
            "queue.html",
            tasks=tasks,
            counts=queue_counts_map(),
            statistics=statistics_rows,
            filter_status=filter_status,
        )

    @app.route("/queue/process")
    def queue_process():
        pending_count = QueryQueue.query.filter(QueryQueue.status.in_(["待处理", "pending"])).count()
        if pending_count == 0:
            return redirect(url_for("queue"))
        return render_template("queue_process.html", pending_count=pending_count)

    @app.route("/employees/add")
    def employee_add_page():
        return render_template("employee_add.html")

    @app.route("/certificates/add")
    def certificate_add_page():
        employees = [emp.to_dict() for emp in Employee.query.order_by(Employee.name.asc()).all()]
        return render_template(
            "certificate_add.html",
            employees=employees,
            certificate=None,
            action_url="/api/certificates",
            form_method="POST",
        )

    @app.route("/certificates")
    def certificates_page():
        search = request.args.get("search", "").strip()
        query = Certificate.query.join(Employee)
        if search:
            query = query.filter(
                or_(
                    Certificate.certificate_type.contains(search),
                    Certificate.operation_item.contains(search),
                    Certificate.cert_number.contains(search),
                    Employee.name.contains(search),
                )
            )
        certificates = [certificate_list_row(cert) for cert in query.order_by(Certificate.created_at.desc()).all()]
        return render_template("certificates.html", certificates=certificates, search=search)

    @app.route("/certificates/<int:cert_id>/edit")
    def certificate_edit_page(cert_id):
        cert = Certificate.query.get_or_404(cert_id)
        cert.sync_alias_fields()
        employees = [emp.to_dict() for emp in Employee.query.order_by(Employee.name.asc()).all()]
        return render_template(
            "certificate_add.html",
            employees=employees,
            certificate=cert.to_dict(),
            action_url=f"/api/certificates/{cert_id}",
            form_method="PUT",
        )

    @app.route("/certificates/<int:cert_id>/delete")
    def certificate_delete_page(cert_id):
        cert = Certificate.query.get_or_404(cert_id)
        db.session.delete(cert)
        db.session.commit()
        return redirect(url_for("certificates_page"))

    @app.route("/export/employees")
    def export_employees():
        wb = Workbook()
        ws = wb.active
        ws.title = "员工列表"
        headers = ["工号", "姓名", "身份证号", "部门", "岗位", "电话", "状态", "证书数"]
        write_excel_header(ws, headers)
        for emp in Employee.query.order_by(Employee.name.asc()).all():
            ws.append(
                [
                    emp.employee_code or "",
                    emp.name,
                    emp.id_number,
                    emp.department or "",
                    emp.position or "",
                    emp.phone or "",
                    emp.employment_status or "",
                    len(emp.certificates),
                ]
            )
        return excel_response(wb, f"员工列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    @app.route("/export/certificates")
    def export_certificates():
        wb = Workbook()
        ws = wb.active
        ws.title = "证书列表"
        headers = ["员工", "身份证号", "证书类型", "作业类别", "操作项目", "证书编号", "发证机关", "有效期至", "复审日期", "状态"]
        write_excel_header(ws, headers)
        for cert in Certificate.query.join(Employee).order_by(Employee.name.asc()).all():
            cert.sync_alias_fields()
            ws.append(
                [
                    cert.employee.name if cert.employee else "",
                    cert.employee.id_number if cert.employee else "",
                    cert.certificate_type,
                    cert.cert_category or "",
                    cert.operation_item or "",
                    cert.cert_number or "",
                    cert.issuing_authority or "",
                    format_date(cert.valid_until),
                    format_date(cert.review_date),
                    evaluate_certificate_status(cert),
                ]
            )
        return excel_response(wb, f"证书列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    @app.route("/api/employees", methods=["POST"])
    def api_add_employee():
        data = parse_payload()
        try:
            employee = create_or_update_employee(data)
            return jsonify({"success": True, "employee_id": employee.id})
        except ValueError as exc:
            return jsonify({"success": False, "error": str(exc)}), 400

    @app.route("/api/employees/list")
    def api_employees_list():
        return jsonify([emp.to_dict() for emp in Employee.query.order_by(Employee.name.asc()).all()])

    @app.route("/api/employees/search")
    def api_employees_search():
        keyword = request.args.get("keyword", "").strip()
        if not keyword:
            return jsonify([])
        rows = [
            employee_search_row(emp)
            for emp in Employee.query.filter(Employee.name.contains(keyword))
            .order_by(Employee.name.asc())
            .limit(20)
            .all()
        ]
        return jsonify(rows)

    @app.route("/api/employees/<int:emp_id>", methods=["GET"])
    def api_get_employee(emp_id):
        return jsonify(Employee.query.get_or_404(emp_id).to_dict())

    @app.route("/api/employees/<int:emp_id>", methods=["PUT"])
    def api_update_employee(emp_id):
        employee = Employee.query.get_or_404(emp_id)
        data = parse_payload()
        apply_employee_changes(employee, data)
        db.session.commit()
        return jsonify({"success": True})

    @app.route("/api/employees/<int:emp_id>", methods=["DELETE"])
    def api_delete_employee(emp_id):
        employee = Employee.query.get_or_404(emp_id)
        db.session.delete(employee)
        db.session.commit()
        return jsonify({"success": True})

    @app.route("/api/certificates", methods=["POST"])
    def api_add_certificate():
        data = parse_payload()
        try:
            cert = create_manual_certificate(data)
            return jsonify({"success": True, "certificate_id": cert.id})
        except ValueError as exc:
            return jsonify({"success": False, "error": str(exc)}), 400

    @app.route("/api/certificates/<int:cert_id>", methods=["GET"])
    def api_get_certificate(cert_id):
        cert = Certificate.query.get_or_404(cert_id)
        return jsonify(cert.to_dict())

    @app.route("/api/certificates/<int:cert_id>", methods=["PUT"])
    def api_update_certificate(cert_id):
        cert = Certificate.query.get_or_404(cert_id)
        data = parse_payload()
        apply_certificate_changes(cert, data)
        db.session.commit()
        return jsonify({"success": True})

    @app.route("/api/certificates/<int:cert_id>", methods=["DELETE"])
    def api_delete_certificate(cert_id):
        cert = Certificate.query.get_or_404(cert_id)
        db.session.delete(cert)
        db.session.commit()
        return jsonify({"success": True})

    @app.route("/api/queue/stats")
    def queue_stats():
        stats = queue_counts_map()
        stats["full_refresh"] = full_refresh_status()
        return jsonify(stats)

    @app.route("/queue/add", methods=["POST"])
    def queue_add():
        employees_text = request.form.get("employees", "")
        query_type = normalize_query_type(request.form.get("query_type", "全部证书"))
        priority = request.form.get("priority", 5, type=int)
        result = add_queue_tasks(employees_text, query_type, priority)
        if request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html:
            return jsonify({"success": True, **result})
        return redirect(url_for("queue"))

    @app.route("/queue/clear", methods=["POST"])
    def queue_clear():
        deleted = QueryQueue.query.filter(QueryQueue.status.in_(["已完成", "失败", "无记录"])).delete(
            synchronize_session=False
        )
        db.session.commit()
        if request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html:
            return jsonify({"success": True, "deleted": deleted})
        return redirect(url_for("queue"))

    @app.route("/api/queue/trigger", methods=["POST"])
    def api_queue_trigger():
        limit = request.get_json(silent=True) or {}
        service = QueryService()
        refresh_result = enqueue_full_refresh_if_due()
        recovered = service.reclaim_stale_processing_tasks()
        pending_count = QueryQueue.query.filter(QueryQueue.status.in_(["待处理", "pending"])).count()
        if pending_count == 0:
            return jsonify(
                {
                    "success": False,
                    "error": "没有待处理任务",
                    "full_refresh": refresh_result,
                    "recovered_count": len(recovered),
                }
            )
        results = service.process_queue(limit=int(limit.get("limit", 10)))
        return jsonify(
            {
                "success": True,
                "full_refresh": refresh_result,
                "pending_count": pending_count,
                "recovered_count": len(recovered),
                "results": results,
            }
        )

    @app.route("/api/queue/full-refresh", methods=["POST"])
    def api_queue_full_refresh():
        payload = request.get_json(silent=True) or {}
        result = enqueue_full_refresh_if_due(
            force=bool(payload.get("force", False)),
            cycle_days=payload.get("cycle_days"),
        )
        return jsonify({"success": True, **result})

    @app.route("/api/queue/process-one", methods=["POST"])
    def process_one():
        data = request.get_json(force=True)
        result = QueryService().query_certificate(data.get("id_number"), data.get("cert_type"), data.get("name"))
        return jsonify(result)

    @app.route("/api/query/direct", methods=["POST"])
    def direct_query():
        data = request.get_json(force=True)
        employee = Employee.query.filter_by(id_number=data["id_number"]).first()
        if not employee:
            employee = create_or_update_employee(
                {"name": data.get("name") or data["id_number"], "id_number": data["id_number"]}
            )
        result = QueryService().query_certificate(employee.id_number, normalize_query_type(data.get("query_type")), employee.name)
        saved = []
        if result.get("success"):
            for cert_data in result.get("data", []):
                save_result = QueryService().save_certificate(employee.id, cert_data, raw_data=result)
                if save_result.get("success"):
                    saved.append(save_result["certificate"])
        return jsonify({"success": result.get("success", False), "result": result, "saved": saved})

    @app.errorhandler(404)
    def not_found(_):
        return render_template("404.html"), 404

    @app.errorhandler(500)
    def internal_error(err):
        logger.exception("Unhandled application error: %s", err)
        return render_template("500.html"), 500


def queue_counts_map():
    rows = (
        db.session.query(QueryQueue.status, func.count(QueryQueue.id))
        .group_by(QueryQueue.status)
        .all()
    )
    base = {"pending": 0, "processing": 0, "completed": 0, "failed": 0}
    label_map = {
        "待处理": "pending",
        "pending": "pending",
        "处理中": "processing",
        "processing": "processing",
        "已完成": "completed",
        "completed": "completed",
        "无记录": "completed",
        "失败": "failed",
        "failed": "failed",
    }
    for status, count in rows:
        key = label_map.get(status)
        if key:
            base[key] += int(count)
    return base


FULL_REFRESH_LAST_ENQUEUED_AT_KEY = "full_refresh_last_enqueued_at"
FULL_REFRESH_LAST_RESULT_KEY = "full_refresh_last_result"
FULL_REFRESH_CYCLE_DAYS_KEY = "full_refresh_cycle_days"


def get_system_config_value(config_key, default=None):
    row = db.session.get(SystemConfig, config_key)
    return row.config_value if row else default


def set_system_config_value(config_key, config_value, description=""):
    row = db.session.get(SystemConfig, config_key)
    if row:
        row.config_value = str(config_value)
        if description:
            row.description = description
    else:
        row = SystemConfig(
            config_key=config_key,
            config_value=str(config_value),
            description=description,
        )
        db.session.add(row)
    return row


def active_employee_query():
    query = Employee.query.filter(Employee.id_number.isnot(None))
    query = query.filter(func.length(func.trim(Employee.id_number)) == 18)
    query = query.filter(
        or_(
            Employee.employment_status.is_(None),
            Employee.employment_status == "",
            Employee.employment_status.in_(["在职", "01", "active", "ACTIVE"]),
        )
    )
    return query


def full_refresh_cycle_days():
    raw_value = get_system_config_value(
        FULL_REFRESH_CYCLE_DAYS_KEY,
        os.environ.get("FULL_REFRESH_CYCLE_DAYS", "7"),
    )
    try:
        return max(1, int(raw_value))
    except (TypeError, ValueError):
        return 7


def full_refresh_status():
    cycle_days = full_refresh_cycle_days()
    last_enqueued_at_raw = get_system_config_value(FULL_REFRESH_LAST_ENQUEUED_AT_KEY)
    last_result_raw = get_system_config_value(FULL_REFRESH_LAST_RESULT_KEY)
    pending_count = QueryQueue.query.filter(QueryQueue.status.in_(["待处理", "pending", "处理中", "processing"])).count()

    last_enqueued_at = None
    next_due_at = None
    due_now = True
    if last_enqueued_at_raw:
        try:
            last_enqueued_at = datetime.fromisoformat(last_enqueued_at_raw)
            next_due_at = last_enqueued_at + timedelta(days=cycle_days)
            due_now = datetime.now() >= next_due_at
        except ValueError:
            last_enqueued_at = None
            next_due_at = None

    status = {
        "cycle_days": cycle_days,
        "last_enqueued_at": format_datetime(last_enqueued_at),
        "next_due_at": format_datetime(next_due_at),
        "due_now": due_now,
        "pending_or_processing_count": pending_count,
    }
    if last_result_raw:
        try:
            status["last_result"] = json.loads(last_result_raw)
        except ValueError:
            status["last_result"] = {"raw": last_result_raw}
    return status


def enqueue_full_refresh_if_due(force=False, cycle_days=None):
    cycle_days = max(1, int(cycle_days or full_refresh_cycle_days()))
    now = datetime.now()
    pending_count = QueryQueue.query.filter(QueryQueue.status.in_(["待处理", "pending", "处理中", "processing"])).count()

    last_enqueued_at_raw = get_system_config_value(FULL_REFRESH_LAST_ENQUEUED_AT_KEY)
    due_now = True
    last_enqueued_at = None
    if last_enqueued_at_raw:
        try:
            last_enqueued_at = datetime.fromisoformat(last_enqueued_at_raw)
            due_now = now >= last_enqueued_at + timedelta(days=cycle_days)
        except ValueError:
            last_enqueued_at = None
            due_now = True

    result = {
        "triggered": False,
        "forced": force,
        "cycle_days": cycle_days,
        "last_enqueued_at": format_datetime(last_enqueued_at),
        "skipped_reason": None,
        "added": 0,
        "existing": 0,
        "eligible": 0,
    }

    if not force and not due_now:
        result["skipped_reason"] = "not_due"
        set_system_config_value(
            FULL_REFRESH_CYCLE_DAYS_KEY,
            cycle_days,
            "全员证书刷新周期（天）",
        )
        set_system_config_value(
            FULL_REFRESH_LAST_RESULT_KEY,
            json.dumps(result, ensure_ascii=False),
            "最近一次全员刷新入队结果",
        )
        db.session.commit()
        return result

    if not force and pending_count > 0:
        result["skipped_reason"] = "queue_busy"
        set_system_config_value(
            FULL_REFRESH_CYCLE_DAYS_KEY,
            cycle_days,
            "全员证书刷新周期（天）",
        )
        set_system_config_value(
            FULL_REFRESH_LAST_RESULT_KEY,
            json.dumps(result, ensure_ascii=False),
            "最近一次全员刷新入队结果",
        )
        db.session.commit()
        return result

    employees = active_employee_query().order_by(Employee.id.asc()).all()
    result["eligible"] = len(employees)

    existing_keys = {
        row[0]
        for row in db.session.query(QueryQueue.id_card)
        .filter(QueryQueue.status.in_(["待处理", "pending", "处理中", "processing"]))
        .all()
    }

    created = 0
    existing = 0
    for employee in employees:
        if employee.id_number in existing_keys:
            existing += 1
            continue

        task = QueryQueue(
            employee_id=employee.id,
            employee_code=employee.employee_code,
            name=employee.name,
            id_card=employee.id_number,
            query_type="全部证书",
            priority=int(os.environ.get("FULL_REFRESH_PRIORITY", "3")),
            status="待处理",
        )
        db.session.add(task)
        existing_keys.add(employee.id_number)
        created += 1

    set_system_config_value(
        FULL_REFRESH_CYCLE_DAYS_KEY,
        cycle_days,
        "全员证书刷新周期（天）",
    )
    set_system_config_value(
        FULL_REFRESH_LAST_ENQUEUED_AT_KEY,
        now.isoformat(timespec="seconds"),
        "最近一次全员证书刷新入队时间",
    )
    result["triggered"] = True
    result["added"] = created
    result["existing"] = existing
    result["last_enqueued_at"] = format_datetime(now)
    db.session.add(
        UpdateNotification(
            title="全员证书刷新已入队",
            content=f"按 {cycle_days} 天周期为 {created} 名员工创建了全员更新任务。",
            notification_type="info",
        )
    )
    set_system_config_value(
        FULL_REFRESH_LAST_RESULT_KEY,
        json.dumps(result, ensure_ascii=False),
        "最近一次全员刷新入队结果",
    )
    db.session.commit()
    return result


def expiring_query(days):
    today = date.today()
    deadline = today.fromordinal(today.toordinal() + max(1, min(days, 365)))
    return (
        Certificate.query.join(Employee)
        .filter(Certificate.valid_until.isnot(None))
        .filter(Certificate.valid_until >= today, Certificate.valid_until <= deadline)
        .order_by(Certificate.valid_until.asc())
    )


def parse_payload():
    if request.is_json:
        return request.get_json() or {}
    if request.form:
        return dict(request.form)
    return {}


def create_or_update_employee(data):
    name = (data.get("name") or "").strip()
    id_number = (data.get("id_number") or data.get("id_card") or "").strip().upper()
    if not name:
        raise ValueError("姓名为必填项")
    if len(id_number) != 18:
        raise ValueError("身份证号必须为18位")

    employee = Employee.query.filter_by(id_number=id_number).first()
    if employee:
        apply_employee_changes(employee, data)
    else:
        employee = Employee(name=name, id_number=id_number)
        apply_employee_changes(employee, data)
        db.session.add(employee)
    db.session.commit()
    return employee


def apply_employee_changes(employee, data):
    employee.name = (data.get("name") or employee.name or "").strip()
    employee.id_number = (data.get("id_number") or data.get("id_card") or employee.id_number or "").strip().upper()
    employee.employee_code = (data.get("employee_code") or employee.employee_code or "").strip()
    employee.department = (data.get("department") or employee.department or "").strip()
    employee.position = (data.get("position") or employee.position or "").strip()
    employee.phone = (data.get("phone") or employee.phone or "").strip()
    employee.employment_status = (data.get("employment_status") or employee.employment_status or "在职").strip()


def create_manual_certificate(data):
    employee_id = int(data.get("employee_id") or 0)
    employee = db.session.get(Employee, employee_id)
    if not employee:
        raise ValueError("员工不存在")

    certificate_type = (data.get("certificate_type") or "").strip()
    if not certificate_type:
        raise ValueError("证书类型为必填项")

    cert = Certificate(employee_id=employee_id)
    apply_certificate_changes(cert, data)
    db.session.add(cert)
    db.session.commit()
    return cert


def apply_certificate_changes(cert, data):
    cert.cert_type = normalize_query_type(data.get("cert_type") or data.get("certificate_type") or cert.cert_type)
    cert.certificate_type = (data.get("certificate_type") or cert.certificate_type or cert.cert_type or "").strip()
    cert.cert_category = (data.get("cert_category") or cert.cert_category or "").strip()
    cert.operation_item = (data.get("operation_item") or cert.operation_item or "").strip()
    cert.cert_item = (data.get("cert_item") or cert.operation_item or cert.cert_item or "").strip()
    cert.cert_name = cert.operation_item or cert.certificate_type
    cert.cert_number = (data.get("cert_number") or cert.cert_number or "").strip()
    cert.issuing_authority = (data.get("issuing_authority") or cert.issuing_authority or "").strip()
    cert.issue_authority = (data.get("issue_authority") or cert.issuing_authority or cert.issue_authority or "").strip()
    cert.issue_date = parse_date(data.get("issue_date")) or cert.issue_date
    cert.valid_from = parse_date(data.get("valid_from")) or cert.valid_from or cert.issue_date
    cert.valid_until = parse_date(data.get("valid_until") or data.get("expiry_date")) or cert.valid_until
    cert.expiry_date = cert.valid_until
    cert.review_date = parse_date(data.get("review_date")) or cert.review_date
    cert.actual_review_date = parse_date(data.get("actual_review_date")) or cert.actual_review_date
    cert.status = evaluate_certificate_status(cert)
    cert.sync_alias_fields()


def parse_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            normalized = str(value).strip()
            parsed = datetime.strptime(normalized, fmt)
            return parsed.date()
        except ValueError:
            continue
    raise ValueError("日期格式无效，应为 YYYY-MM-DD")


def employee_search_row(emp):
    row = emp.to_dict()
    row["id_card"] = row["id_number"]
    return row


def detail_certificate_row(cert):
    cert.sync_alias_fields()
    expiry = cert.valid_until
    days = (expiry - date.today()).days if expiry else None
    status = evaluate_certificate_status(cert)
    return {
        "id": cert.id,
        "certificate_type": cert.certificate_type,
        "operation_item": cert.operation_item,
        "issuing_authority": cert.issuing_authority,
        "initial_issue_date": format_date(cert.issue_date),
        "expiry_date": format_date(expiry),
        "review_date": format_date(cert.review_date),
        "days_until_expiry": days,
        "status": status,
        "status_class": status_class(status),
    }


def certificate_row(cert):
    cert.sync_alias_fields()
    expiry = cert.valid_until
    days_to_expiry = (expiry - date.today()).days if expiry else None
    days_to_review = (cert.review_date - date.today()).days if cert.review_date else None
    status = evaluate_certificate_status(cert)
    display_status = status
    if status == "待复审":
        display_status = "即将到期"
    return {
        "employee_id": cert.employee_id,
        "certificate_id": cert.id,
        "name": cert.employee.name if cert.employee else "",
        "id_card": cert.employee.id_number if cert.employee else "",
        "certificate_type": cert.certificate_type,
        "operation_item": cert.operation_item,
        "issuing_authority": cert.issuing_authority,
        "review_date": format_date(cert.review_date),
        "actual_review_date": format_date(cert.actual_review_date),
        "expiry_date": format_date(expiry),
        "days_until_expiry": days_to_expiry,
        "days_to_review": days_to_review,
        "display_status": display_status,
        "status_class": status_class(display_status),
    }


def expiring_row(cert):
    cert.sync_alias_fields()
    days = (cert.valid_until - date.today()).days if cert.valid_until else None
    status = evaluate_certificate_status(cert)
    return {
        "name": cert.employee.name if cert.employee else "",
        "phone": cert.employee.phone if cert.employee else "",
        "department": cert.employee.department if cert.employee else "",
        "certificate_type": cert.certificate_type,
        "operation_item": cert.operation_item,
        "expiry_date": format_date(cert.valid_until),
        "days_until_expiry": days,
        "status": status,
        "status_class": status_class(status),
    }


def certificate_list_row(cert):
    cert.sync_alias_fields()
    return {
        "id": cert.id,
        "employee_name": cert.employee.name if cert.employee else "",
        "certificate_type": cert.certificate_type,
        "operation_item": cert.operation_item,
        "expiry_date": format_date(cert.valid_until),
        "status": evaluate_certificate_status(cert),
    }


def status_class(status):
    if status in {"有效"}:
        return "status-valid"
    if status in {"已过期", "失败"}:
        return "status-expired"
    if status in {"即将到期", "待复审", "无记录"}:
        return "status-warning"
    return "status-notice"


def notification_to_dict(item):
    return {
        "id": item.id,
        "title": item.title,
        "content": item.content,
        "notification_type": item.notification_type,
        "created_at": format_datetime(item.created_at),
    }


def add_queue_tasks(employees_text, query_type, priority):
    added = 0
    duplicate = 0
    not_found = 0

    for raw_line in employees_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        parts = line.split()
        employee = None
        employee_code = None
        name = None

        if len(parts) == 1:
            token = parts[0]
            if len(token) == 18:
                employee = Employee.query.filter_by(id_number=token.upper()).first()
            else:
                employee = Employee.query.filter_by(name=token).first()
        else:
            employee_code = parts[0]
            name = parts[-1]
            employee = Employee.query.filter(
                or_(Employee.employee_code == employee_code, Employee.name == name)
            ).first()

        if not employee:
            not_found += 1
            continue

        exists = QueryQueue.query.filter(
            QueryQueue.id_card == employee.id_number,
            QueryQueue.query_type == query_type,
            QueryQueue.status.in_(["待处理", "处理中"]),
        ).first()
        if exists:
            duplicate += 1
            continue

        task = QueryQueue(
            employee_id=employee.id,
            employee_code=employee.employee_code or employee_code,
            name=employee.name,
            id_card=employee.id_number,
            query_type=query_type,
            priority=priority,
            status="待处理",
        )
        db.session.add(task)
        added += 1

    db.session.commit()
    return {
        "added": added,
        "duplicate": duplicate,
        "not_found": not_found,
        "message": f"添加完成: 成功 {added} 条, 重复 {duplicate} 条, 未找到 {not_found} 条",
    }


def write_excel_header(ws, headers):
    ws.append(headers)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def excel_response(wb, filename):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


app = create_app()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
